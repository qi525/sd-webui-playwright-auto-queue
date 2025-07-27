# my_tools.py (集成日志配置、文件系统操作、自动打开功能、Excel 辅助功能和应用程序运行器)
from loguru import logger
from pathlib import Path
import os
import datetime
import sys
import time
import subprocess
import urllib.parse # 新增：用于 URL 解析，支持 get_url_host, get_domain_from_url
import re # 新增：用于正则表达式，支持 clean_filename_string
import hashlib # 新增：用于哈希计算，支持 generate_unique_filename
import shutil # 新增：用于文件复制，支持 copy_file_robustly
import asyncio # 新增：用于 run_application 中的 asyncio.run
from typing import Tuple, List, Optional, Dict, Any, Callable # 导入 Any, Callable

# openpyxl 相关导入
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# 获取当前脚本的父目录（项目根目录）
BASE_DIR = Path(__file__).resolve().parent

# --- 从 file_system_utils.py 复制过来的函数 ---

def normalize_drive_letter(path_str: str) -> str:
    """
    如果路径以驱动器号开头，将其转换为大写。
    例如: c:\\test -> C:\\test
    """
    if sys.platform.startswith('win') and len(path_str) >= 2 and path_str[1] == ':':
        return path_str[0].upper() + path_str[1:]
    return path_str

def create_directory_if_not_exists(directory_path: Path, logger_obj) -> bool:
    """
    如果指定目录不存在，则创建它。
    Args:
        directory_path (Path): 要创建的目录路径。
        logger_obj: 日志管理器实例。
    Returns:
        bool: 如果目录存在或成功创建，则返回True；否则返回False。
    """
    if not directory_path.exists():
        try:
            os.makedirs(directory_path)
            if logger_obj:
                logger_obj.info(f"已创建目录: {normalize_drive_letter(str(directory_path))}")
            return True
        except OSError as e:
            if logger_obj:
                logger_obj.error(f"创建目录失败 {normalize_drive_letter(str(directory_path))}: {e}")
            return False
    return True

def copy_file(source_path: Path, destination_path: Path, logger_obj) -> bool:
    """
    复制文件从源路径到目标路径。
    增加对权限错误的捕获和提示。
    """
    if not source_path.exists():
        if logger_obj:
            logger_obj.error(f"错误: 源文件不存在，无法复制: {normalize_drive_letter(str(source_path))}")#error
        return False

    try:
        shutil.copy2(str(source_path), str(destination_path)) 
        if logger_obj:
            logger_obj.info(f"已复制 '{normalize_drive_letter(str(source_path))}' 到 '{normalize_drive_letter(str(destination_path))}'")
        return True
    except PermissionError as e:
        if logger_obj:
            logger_obj.critical(f"权限错误: 复制文件从 '{normalize_drive_letter(str(source_path))}' 到 '{normalize_drive_letter(str(destination_path))}' 失败: {e}. 请确保目标文件未被其他程序（如Excel）占用。")
        return False
    except Exception as e:
        if logger_obj:
            logger_obj.error(f"错误: 复制文件从 '{normalize_drive_letter(str(source_path))}' 到 '{normalize_drive_letter(str(destination_path))}' 失败: {e}")#error
        return False


# --- 从 file_opener.py 复制过来的函数 ---

def open_output_files_automatically(
    file_paths: List[Path],
    logger_obj
):
    """
    根据用户设置自动打开生成的输出文件（Excel和Log文件）。
    Args:
        file_paths (List[Path]): 包含要打开的文件路径的列表。
        logger_obj (logger): Loguru logger 实例。
    """
    if os.getenv("DISABLE_AUTO_OPEN", "0") == "1":
        logger_obj.info("已禁用自动打开文件功能。")
        return

    # 定义延迟时间 (秒)
    OPEN_FILE_DELAY_SECONDS = 2

    for file_path in file_paths:
        logger_obj.debug(f"尝试打开文件 '{normalize_drive_letter(str(file_path))}' 前，等待 {OPEN_FILE_DELAY_SECONDS} 秒。")
        time.sleep(OPEN_FILE_DELAY_SECONDS)

        actual_path_to_open = file_path
        # 特殊处理 Loguru 压缩后的日志文件
        if file_path.suffix == '.log' and not file_path.exists():
            zip_path = file_path.with_suffix('.zip')
            if zip_path.exists():
                actual_path_to_open = zip_path
                logger_obj.info(f"日志文件 '{normalize_drive_letter(str(file_path))}' 不存在，尝试打开压缩文件: {normalize_drive_letter(str(zip_path))}")

        if not actual_path_to_open.exists():
            logger_obj.warning(f"警告: 无法自动打开文件 '{normalize_drive_letter(str(actual_path_to_open))}'，因为文件不存在。")
            continue

        try:
            normalized_path = normalize_drive_letter(str(actual_path_to_open))
            logger_obj.info(f"自动打开: {normalized_path}")
            if sys.platform == "win32":
                os.startfile(normalized_path)
            elif sys.platform == "darwin": # macOS
                subprocess.run(['open', normalized_path], check=True)
            else: # Linux
                subprocess.run(['xdg-open', normalized_path], check=True)
        except FileNotFoundError:
            logger_obj.error(f"错误: 无法找到打开文件 '{normalized_path}' 的应用程序。请手动打开。", exc_info=True)
        except Exception as e:
            logger_obj.error(f"自动打开文件 '{normalized_path}' 时发生意外错误: {e}", exc_info=True)

# --- setup_logger 函数，现在是 my_tools 的核心日志配置部分 ---

def setup_logger(log_base_directory: Path = BASE_DIR / "logs") -> Tuple[Path, Path]:
    """
    配置 Loguru 日志器，设置多个日志输出目标（主日志、错误/警告日志、控制台）。
    此函数应在程序启动时调用一次。
    主日志文件将以 DEBUG 级别记录所有详细信息。

    Args:
        log_base_directory (Path): 日志文件的根目录。默认为脚本运行目录下的 "logs" 文件夹。
        # auto_open (bool): 此参数已移除，日志文件不再在 setup_logger 中自动打开。

    Returns:
        Tuple[Path, Path]: 返回一个元组，包含 (error_warning_log_path, main_log_path)。
    """
    logger.remove()  # 移除所有默认或之前添加的处理器

    # 使用内部的 create_directory_if_not_exists 函数
    if not create_directory_if_not_exists(log_base_directory, logger):
        logger.critical(f"关键错误: 无法创建日志文件夹 {log_base_directory}. 日志将仅打印到控制台.")
        return Path("invalid_log_path_error"), Path("invalid_log_path_main")

    # 1. 配置主程序日志 (DEBUG 及以上，到独立文件，包含所有详细信息)
    main_log_file_name = f"main_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    main_log_path = log_base_directory / main_log_file_name

    logger.add(
        sink=main_log_path,
        level="DEBUG",
        format="{time:YYYY-MM-DD HH:mm:ss} [{level}] <cyan>{file}</cyan>:<cyan>{function}</cyan>:<cyan>{line}</cyan> - {message}",
        rotation="10 MB",
        retention="7 days",
        compression="zip",
        enqueue=True,
        backtrace=True,
        diagnose=True
    )

    # 2. 配置错误和警告日志 (WARNING 及以上，到独立文件)
    error_warning_log_file_name = f"error_warning_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    error_warning_log_path = log_base_directory / error_warning_log_file_name

    logger.add(
        sink=error_warning_log_path,
        level="WARNING",
        format="{time:YYYY-MM-DD HH:mm:ss} [{level}] <red>File \"{file.path}\", line {line}, in {function}</red> - {message}",
        rotation="10 MB",
        retention="7 days",
        compression="zip",
        enqueue=True,
        backtrace=True,
        diagnose=True,
        filter=lambda record: record["level"].name in ["WARNING", "ERROR", "CRITICAL"]
    )

    # 3. 配置控制台输出（仅 INFO 及以上，保持简洁）
    logger.add(
        sys.stderr,
        level="INFO",
        format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> <level>[{level}]</level> <yellow>File \"{file.path}\", line {line}, in {function}</yellow> - <level>{message}</level>",
        colorize=True,
        filter=lambda record: record["level"].name in ["INFO", "SUCCESS"]
    )

    return error_warning_log_path, main_log_path

# --- 新增函数：在程序结束时自动打开日志文件 ---
async def open_completed_logs( # 函数名已更新
    main_log_path: Path,
    error_log_path: Path,
    logger_obj,
    is_auto_open: bool,
    extra_files: Optional[List[Path]] = None # 新增参数用于汇总Excel
):
    """
    在程序完成运行后检查并自动打开主日志和错误日志文件。
    错误日志文件仅在非空时打开。

    Args:
        main_log_path (Path): 主日志文件的路径。
        error_log_path (Path): 错误日志文件的路径。
        logger_obj: 日志管理器实例。
        is_auto_open (bool): 是否启用自动打开功能。
        extra_files (Optional[List[Path]]): 除了日志文件外，额外的需要自动打开的文件列表。
    """
    if not is_auto_open:
        logger_obj.info("已禁用自动打开日志文件功能。")
        return

    files_to_open = []

    # 总是尝试打开主日志文件（如果存在）
    if main_log_path and main_log_path.exists():
        files_to_open.append(main_log_path)
        logger_obj.info(f"主日志文件 '{normalize_drive_letter(str(main_log_path.name))}' 将在程序结束时尝试打开。")
    else:
        logger_obj.warning(f"主日志文件 '{normalize_drive_letter(str(main_log_path))}' 不存在，无法打开。")

    # 错误日志文件仅在非空时打开
    if error_log_path and error_log_path.exists():
        if error_log_path.stat().st_size > 0:
            files_to_open.append(error_log_path)
            logger_obj.info(f"错误/警告日志文件 '{normalize_drive_letter(str(error_log_path.name))}' 非空，将在程序结束时尝试打开。")
        else:
            logger_obj.info(f"错误/警告日志文件 '{normalize_drive_letter(str(error_log_path.name))}' 为空 (0KB)，无需自动打开。")
    else:
        logger_obj.info(f"错误/警告日志文件 '{normalize_drive_letter(str(error_log_path.name))}' 不存在，无需自动打开。")

    if extra_files:
        for file_path in extra_files:
            if file_path and file_path.exists():
                files_to_open.append(file_path)
                logger_obj.info(f"额外文件 '{normalize_drive_letter(str(file_path.name))}' 将在程序结束时尝试打开。")
            else:
                logger_obj.warning(f"额外文件 '{normalize_drive_letter(str(file_path))}' 不存在，无法打开。")

    if files_to_open:
        # 使用 asyncio.to_thread 在事件循环中运行同步的 open_output_files_automatically
        await asyncio.to_thread(open_output_files_automatically, files_to_open, logger_obj)
    else:
        logger_obj.info("没有日志文件需要自动打开。")


# --- 原 open_error_log_if_not_empty 函数已移除，其功能已整合到 open_completed_logs 中 ---

# --- 新增类：运行状态管理器 ---
class RunStatusManager:
    """
    负责管理应用程序的运行状态，主要通过检查错误日志文件来确定。
    """
    def __init__(self, error_log_path: Path, logger_obj):
        """
        初始化 RunStatusManager。
        Args:
            error_log_path (Path): 错误日志文件的路径。
            logger_obj: 日志管理器实例。
        """
        self.error_log_path = error_log_path
        self.logger_obj = logger_obj
        self.initial_status_checked = False # 标记是否已进行初始状态检查

    def get_status(self) -> str:
        """
        根据错误日志文件内容确定应用程序的运行状态。
        Returns:
            str: 运行状态（"SUCCESS", "Run Warning", "Run Error", "Invalid Log Path", "Log File Not Found"）。
        """
        run_status = "SUCCESS"
        if self.error_log_path and self.error_log_path.exists():
            try:
                with open(self.error_log_path, 'r', encoding='utf-8') as f:
                    log_content = f.read()
                if "ERROR" in log_content or "CRITICAL" in log_content:
                    run_status = "Run Error"
                elif "WARNING" in log_content:
                    run_status = "Run Warning"
            except Exception as e:
                self.logger_obj.error(f"错误: 读取日志文件 {normalize_drive_letter(str(self.error_log_path))} 时发生错误: {e}")
                run_status = "Run Error"
        elif self.error_log_path is None:
            run_status = "Invalid Log Path"
            self.logger_obj.error("无效的日志文件路径，无法检查运行状态或添加到历史记录。")
        else:
            run_status = "Log File Not Found"
            self.logger_obj.error(f"日志文件 '{normalize_drive_letter(str(self.error_log_path))}' 不存在，无法检查运行状态。")

        if not self.initial_status_checked:
            self.logger_obj.info(f"应用程序初始运行状态（基于日志文件检查）：{run_status}")
            self.initial_status_checked = True
        return run_status


# --- 从 excel_utilities.py 复制过来的常量和函数 ---

HYPERLINK_FONT = Font(color="0000FF", underline="single")

# 定义固定列宽（以字符为单位）
FIXED_COLUMN_WIDTH = 20

# --- Excel Utilities ---

def create_empty_workbook() -> Workbook:
    """
    创建一个空的Excel工作簿，并移除默认创建的Sheet。
    """
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    return wb

def create_sheet_with_headers(
    workbook: Workbook,
    sheet_name: str,
    headers: List[str],
    index: Optional[int] = None
) -> Worksheet:
    """
    在给定的工作簿中创建一个新工作表，并设置其标题行。
    Args:
        workbook (Workbook): openpyxl 工作簿对象。
        sheet_name (str): 新工作表的名称。
        headers (List[str]): 包含所有列标题的列表。
        index (Optional[int]): 工作表插入的位置索引。如果为 None，则添加到最后。
    Returns:
        Worksheet: 新创建的工作表对象。
    """
    ws = workbook.create_sheet(sheet_name, index)
    ws.append(headers)
    return ws

def set_hyperlink_and_style(
    cell,
    location: Optional[str],
    display_text: str,
    logger_obj,
    source_description: str = "未知来源"
) -> bool:
    """
    为openpyxl单元格设置超链接和样式。
    Args:
        cell: openpyxl单元格对象。
        location (Optional[str]): 超链接的目标路径，如果为None或空字符串则不设置超链接。
        display_text (str): 单元格显示的文本。
        logger_obj: 日志管理器实例。
        source_description (str): 用于日志记录的来源描述，方便定位问题。
    Returns:
        bool: 如果成功设置超链接则返回True，否则返回False。
    """
    cell.value = display_text
    try:
        if location:
            cell.hyperlink = location
            cell.font = HYPERLINK_FONT
            return True
        else:
            cell.hyperlink = None
            cell.font = Font(color="000000")
            return False
    except Exception as e:
        logger_obj.error(
            f"错误: 无法为单元格设置超链接或样式 for '{display_text}' (Location: '{location}', Source: {source_description}). 错误: {e}"
        )
        cell.value = display_text
        return False

def set_column_widths(
    worksheet: Worksheet,
    column_widths: Optional[Dict[str, int]] = None,
    default_width: Optional[int] = None,
    logger_obj=None
):
    """
    为给定工作表的列设置宽度。
    可以为所有列设置一个默认宽度，或者为特定列设置指定宽度。
    Args:
        worksheet (Worksheet): openpyxl 工作表对象。
        column_widths (Optional[Dict[str, int]]): 字典，键为列字母（如'A'）或列索引（从1开始），值为宽度。
        default_width (Optional[int]): 如果指定，所有未在 column_widths 中明确设置的列将使用此默认宽度。
        logger_obj (logger): 日志管理器实例。
    """
    if logger_obj is None:
        class DummyLogger:
            def debug(self, msg): pass
            def info(self, msg): pass
            def warning(self, msg): pass
            def error(self, msg): pass
            def critical(self, msg): pass
        logger_obj = DummyLogger()

    try:
        if default_width is not None:
            for col_idx in range(1, worksheet.max_column + 1):
                column_letter = get_column_letter(col_idx)
                if column_widths is None or (column_letter not in column_widths and col_idx not in column_widths):
                    # 显式转换为浮点数
                    worksheet.column_dimensions[column_letter].width = float(default_width)
        if column_widths:
            for key, width in column_widths.items():
                if isinstance(key, str) and len(key) == 1 and key.isalpha():
                    col_letter = key.upper()
                elif isinstance(key, int):
                    col_letter = get_column_letter(key)
                else:
                    logger_obj.warning(f"警告: 无效的列宽度键 '{key}'，请使用列字母（如'A'）或列索引（如1）。")
                    continue
                # 显式转换为浮点数
                worksheet.column_dimensions[col_letter].width = float(width)
        logger_obj.debug(f"已为工作表 '{worksheet.title}' 设置列宽。")
    except Exception as e:
        logger_obj.error(f"错误: 无法为工作表 '{worksheet.title}' 设置列宽。错误: {e}")

def set_fixed_column_widths(worksheet: Worksheet, width: int, logger_obj):
    """
    为给定工作表的所有列设置固定宽度。
    这是一个兼容旧接口的函数，内部调用 set_column_widths。
    Args:
        worksheet (Worksheet): openpyxl 工作表对象。
        width (int): 要设置的列宽。
        logger_obj (logger): 日志管理器实例。
    """
    set_column_widths(worksheet, default_width=width, logger_obj=logger_obj)


# --- 从 app_runner.py 融合进来的函数 ---
def run_application(main_func: Callable[[Dict], Any], app_config: Dict):
    """
    运行应用程序的主入口点，处理异步执行和顶级异常。

    Args:
        main_func (Callable[[Dict], Any]): 应用程序的主异步函数，
                                            它将接收一个配置字典作为参数。
        app_config (Dict): 应用程序的配置字典。
    """
    try:
        asyncio.run(main_func(app_config))
    except KeyboardInterrupt:
        logger.info("用户中断了脚本。")
    except Exception as e:
        logger.critical(f"应用程序执行期间发生未处理的错误: {e}", exc_info=True)


# --- 新增函数：生成带时间后缀的文件名 ---
def generate_timestamped_filename(original_filename: str, timestamp_format: str = "%Y%m%d_%H%M%S") -> str:
    """
    为文件名添加当前时间戳后缀。
    例如: "report.xlsx" -> "report_20240718_153045.xlsx"

    Args:
        original_filename (str): 原始文件名 (包含扩展名)。
        timestamp_format (str): 时间戳的格式字符串，默认为 "年月日_时分秒"。

    Returns:
        str: 添加时间戳后的新文件名。
    """
    name, ext = os.path.splitext(original_filename)
    timestamp = datetime.datetime.now().strftime(timestamp_format)
    return f"{name}_{timestamp}{ext}"


# --- 从 history_execution.py 移动过来的函数：等待文件释放 ---
def wait_for_file_release(file_path: Path, action_description: str, logger_obj, max_attempts: int = 0) -> bool:
    """
    等待文件被释放，如果文件被占用则提示用户关闭。
    此函数现在是一个通用工具函数，可用于任何文件。
    Args:
        file_path (Path): 要检查的文件路径。
        action_description (str): 正在尝试进行的操作描述 (例如 "删除旧的记录文件")。
        logger_obj: 日志管理器实例。
        max_attempts (int): 最大重试次数。如果为0或负数，表示无限次重试。
    Returns:
        bool: 如果文件最终可以访问则返回 True，否则返回 False。
    """
    attempt = 0
    while True:
        attempt += 1
        if max_attempts > 0 and attempt > max_attempts:
            logger_obj.critical(f"致命错误: 文件 '{normalize_drive_letter(str(file_path))}' 经过 {max_attempts} 次尝试后仍被占用，无法进行 '{action_description}' 操作。")
            return False
        
        try:
            # 尝试以独占写入模式打开文件，如果成功则表示文件未被占用
            # 使用 'a' 模式（追加模式）并立即关闭，以测试写入权限而不实际修改文件内容
            with open(file_path, 'a', encoding='utf-8') as f:
                pass
            logger_obj.info(f"文件 '{normalize_drive_letter(str(file_path))}' 已释放，可以继续操作。")
            return True
        except PermissionError as e:
            logger_obj.info(
                #从warning改为info，避免导致history模块最终，结果显示报错
                f"警告: 文件 '{normalize_drive_letter(str(file_path))}' 被占用，无法进行 '{action_description}' 操作。 "
                f"请关闭任何正在使用此文件的程序 (例如 Excel)。 (尝试 {attempt}{'/' + str(max_attempts) if max_attempts > 0 else ''})"
            )
            print(f"\n文件 '{normalize_drive_letter(str(file_path))}' 被占用，无法进行 '{action_description}' 操作。")
            print("请关闭任何正在使用此文件的程序 (例如 Excel)，然后按 Enter 键继续...")
            input() # 等待用户按下回车
        except FileNotFoundError:
            logger_obj.info(f"文件 '{normalize_drive_letter(str(file_path))}' 不存在，无需等待释放。")
            return True # 文件不存在，可以直接进行操作
        except Exception as e:
            logger_obj.error(f"检查文件 '{normalize_drive_letter(str(file_path))}' 访问性时发生意外错误: {e}", exc_info=True)
            return False # 发生其他错误，直接失败

# --- 从 general_image_scarpy_demo.py 移动过来的通用函数 (难度系数 1/10) ---

def get_url_host(url: str) -> str:
    """
    从给定的 URL 中提取网络位置（host:port），即主机名和端口号（如果存在）。
    此函数返回的是原始的网络位置，不进行 'www.' 或端口的移除。
    难度系数: 1/10

    Args:
        url (str): 要解析的完整 URL 字符串。

    Returns:
        str: 提取出的网络位置字符串，如果解析失败则返回空字符串。
    """
    try:
        parsed_url = urllib.parse.urlparse(url)
        return parsed_url.netloc
    except Exception as e:
        logger.warning(f"无法从URL '{url}' 提取主机: {e}。将返回空字符串。")
        return ""

def get_domain_from_url(url: str) -> str:
    """
    从给定的 URL 中提取主域名（不包含 'www.' 前缀和端口号）。
    例如:
    - https://www.google.com/search?q=test -> google.com
    - https://danbooru.donmai.us/ -> danbooru.donmai.us
    难度系数: 1/10

    Args:
        url (str): 要解析的完整 URL 字符串。

    Returns:
        str: 提取出的主域名，如果解析失败则返回 'unknown_domain'。
    """
    try:
        # 使用 urllib.parse.urlparse 解析 URL，将其分解为各个组成部分。
        parsed_url = urllib.parse.urlparse(url)
        # 获取网络位置（通常是域名和端口）
        netloc = parsed_url.netloc
        # 如果网络位置包含端口号（如 example.com:8080），则移除端口号。
        if ':' in netloc:
            netloc = netloc.split(':')[0]
        # 如果域名以 'www.' 开头，则移除 'www.' 前缀，获取更简洁的主域名。
        if netloc.startswith('www.'):
            netloc = netloc[4:]
        return netloc
    except Exception as e:
        # 如果在提取域名过程中发生任何异常，记录警告并返回一个默认域名。
        logger.warning(f"无法从URL '{url}' 提取域名: {e}。将使用 'unknown_domain'。")
        return "unknown_domain"

def clean_filename_string(filename_str: str) -> str:
    """
    清理文件名字符串，移除其中不允许用于文件名的特殊字符，并替换为下划线。
    难度系数: 1/10

    Args:
        filename_str (str): 待清理的文件名字符串。

    Returns:
        str: 清理后的文件名字符串。
    """
    # 移除文件名中所有无效的字符，并替换为下划线。
    return re.sub(r'[\\/:*?"<>|]', '_', filename_str)

def generate_unique_filename(original_filename_stem: str, actual_file_extension: str, full_image_url: str, index: int) -> str:
    """
    根据给定的基础名称、文件扩展名、完整 URL 和索引，生成一个唯一的文件名。
    此函数会处理通用名称，并包含 URL 的哈希值以增加唯一性。
    难度系数: 1/10

    Args:
        original_filename_stem (str): 图片的原始文件名主体（不含扩展名）。
        actual_file_extension (str): 图片的实际文件扩展名（带点，如 '.jpg'）。
        full_image_url (str): 完整的图片 URL，用于生成哈希值。
        index (int): 当前图片在列表中的索引，用于进一步增加唯一性。

    Returns:
        str: 生成的唯一文件名。
    """
    # 将MD5哈希的十六进制字符串转换为整数，然后取绝对值并进行模运算
    url_hash = int(hashlib.md5(full_image_url.encode()).hexdigest(), 16) % 1000000 
    final_filename_base = original_filename_stem
    
    # 处理通用名称：如果原始文件名主体过于通用，可以给它一个更具描述性的前缀。
    if final_filename_base.lower() in ["image", "index", "home", "default", "search", "result", ""]:
        final_filename_base = f"downloaded_image" 

    # 最终的文件名格式：文件名主体_哈希值_索引.扩展名。
    return f"{final_filename_base}_{url_hash}_{index}{actual_file_extension}"

def copy_file_robustly(source_path: Path, destination_path: Path, logger_obj) -> bool:
    """
    尝试复制文件。
    难度系数: 1/10

    Args:
        source_path (Path): 源文件路径。
        destination_path (Path): 目标文件路径。
        logger_obj: 日志管理器实例。

    Returns:
        bool: 如果文件成功复制到目标位置，则返回 True；否则返回 False。
    """
    try:
        shutil.copy2(source_path, destination_path) # 尝试复制文件，保留元数据。
        logger_obj.info(f"Copied file: {source_path.name} -> {destination_path}")
        return True
    except Exception as e:
        logger_obj.error(f"复制文件 {source_path.name} 到 {destination_path} 失败: {e}", exc_info=True)
        return False
